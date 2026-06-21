"""Exportación del informe SAVT a Excel (.xlsx)."""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from savt import __app_name__, __version__
from savt.models import AuditReport
from savt.report_builder import display_title, findings_dataframe_rows, gravity_label
from savt.section_audit import section_audit_summary_rows
from savt.taxonomy import enrich_finding, severity_label


HEADER_FILL = PatternFill("solid", fgColor="06492F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="06492F")


def _autosize_columns(ws, max_width: int = 48) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        length = 0
        for cell in column_cells:
            if cell.value is not None:
                length = max(length, min(len(str(cell.value)), max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, length + 2)


def _write_sheet_from_rows(ws, rows: list[dict], title: str) -> None:
    ws.append([title])
    ws["A1"].font = TITLE_FONT
    ws.append([])

    if not rows:
        ws.append(["Sin datos"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    for row_idx in range(4, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    _autosize_columns(ws)


def build_report_xlsx(report: AuditReport, dashboard: dict) -> bytes:
    """Genera un libro Excel con resumen, apartados, profundidad, checklist y hallazgos."""
    wb = Workbook()
    wb.remove(wb.active)

    filename = report.filename
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    # --- Resumen ---
    ws_summary = wb.create_sheet("Resumen")
    checklist = dashboard.get("checklist") or {}
    summary_rows = [
        {"Campo": "Documento", "Valor": filename},
        {"Campo": "Generado", "Valor": now},
        {"Campo": "Sistema", "Valor": f"{__app_name__} v{__version__}"},
        {"Campo": "Perfil", "Valor": dashboard.get("profile_label", "—")},
        {"Campo": "ICAI", "Valor": f"{dashboard.get('icai', 0)}/100"},
        {"Campo": "Interpretación ICAI", "Valor": dashboard.get("icai_interpretation", "—")},
        {"Campo": "Estado general", "Valor": dashboard.get("readiness", "—")},
        {"Campo": "Motivo principal", "Valor": dashboard.get("main_reason", "—")},
        {"Campo": "Checklist", "Valor": checklist.get("status", "—")},
        {"Campo": "Palabras (cuerpo)", "Valor": report.word_count},
        {"Campo": "Referencias", "Valor": len(report.bibliography)},
        {"Campo": "Errores críticos", "Valor": dashboard.get("errors", 0)},
        {"Campo": "Advertencias", "Valor": dashboard.get("warnings", 0)},
    ]
    _write_sheet_from_rows(ws_summary, summary_rows, "Resumen general SAVT")

    # --- Apartados detectados ---
    detected = dashboard.get("detected_sections") or []
    ws_detected = wb.create_sheet("Apartados detectados")
    detected_rows = [
        {
            "Orden": s.get("order", ""),
            "Apartado canónico": s.get("title", ""),
            "Detectado como": s.get("detected_as", ""),
            "Palabras": s.get("words", 0),
            "% del cuerpo": s.get("percent_label", ""),
        }
        for s in detected
    ]
    _write_sheet_from_rows(ws_detected, detected_rows, "Apartados detectados en el documento")

    # --- Auditoría por apartado ---
    section_audits = dashboard.get("section_audits") or []
    ws_sections = wb.create_sheet("Auditoría por apartado")
    _write_sheet_from_rows(
        ws_sections,
        section_audit_summary_rows(section_audits),
        "Revisión apartado por apartado",
    )

    recon = dashboard.get("citation_reconciliation") or {}
    if recon.get("reconciliation_rows"):
        ws_recon = wb.create_sheet("Cuadre citas")
        recon_rows = list(recon["reconciliation_rows"])
        for note in recon.get("notes") or []:
            recon_rows.append({"Apartado": "Nota", "Apariciones cita": note, "N° refs distintos": ""})
        _write_sheet_from_rows(ws_recon, recon_rows, "Cuadre de citas y referencias")

    # --- Profundidad académica ---
    content = dashboard.get("content_dashboard") or {}
    depth_rows = []
    for item in content.get("section_depth") or []:
        if item.get("depth_status") == "missing" and item.get("words", 0) <= 0:
            continue
        depth_rows.append(
            {
                "Apartado": item.get("title", ""),
                "Detectado como": item.get("detected_as", ""),
                "Palabras": item.get("words", 0),
                "Citas": item.get("citation_count", 0),
                "Densidad citas/100 pal.": item.get("citation_density", 0),
                "Marcadores críticos": item.get("critical_markers", 0),
                "Ind. hallazgos": item.get("result_markers", 0) or "—",
                "Profundidad": item.get("depth_label", ""),
                "Motivo": item.get("depth_reason", ""),
            }
        )
    ws_depth = wb.create_sheet("Profundidad académica")
    _write_sheet_from_rows(ws_depth, depth_rows, "Profundidad académica por apartado")

    # --- Checklist ---
    checklist_rows = [
        {
            "Apartado": item.get("label", ""),
            "Estado": "Conforme" if item.get("ok") else ("Parcial" if item.get("partial") else "Revisar"),
        }
        for item in (checklist.get("items") or [])
    ]
    ws_check = wb.create_sheet("Checklist")
    _write_sheet_from_rows(ws_check, checklist_rows, "Checklist de presentación")

    # --- Revisión por capítulo ---
    chapter_rows = []
    for review in dashboard.get("chapter_reviews") or []:
        chapter_rows.append(
            {
                "Apartado": review.get("title", ""),
                "Estado": review.get("status", ""),
                "Conforme": "Sí" if review.get("ok") else "No",
                "Parcial": "Sí" if review.get("partial") else "No",
                "Resumen": review.get("summary", ""),
                "Por qué importa": review.get("why", ""),
                "Cómo corregir": review.get("how_to_fix", ""),
            }
        )
    ws_chapters = wb.create_sheet("Revisión por capítulo")
    _write_sheet_from_rows(ws_chapters, chapter_rows, "Revisión detallada por capítulo")

    # --- Hallazgos ---
    ws_findings = wb.create_sheet("Hallazgos")
    _write_sheet_from_rows(ws_findings, findings_dataframe_rows(report), "Hallazgos y recomendaciones")

    # --- Bibliografía ---
    bib_rows = []
    style = report.metadata.get("citation_style", "numbered")
    for num in sorted(report.bibliography):
        ref = report.bibliography[num]
        if style == "apa":
            cited = "Sí" if ref.key in report.cited_keys else "Parcial/No"
        else:
            cited = "Sí" if num in report.cited_numbers else "No"
        bib_rows.append(
            {
                "N°": num,
                "Citada": cited,
                "Año": ref.year,
                "DOI": ref.doi,
                "Referencia": ref.raw[:300],
            }
        )
    ws_bib = wb.create_sheet("Bibliografía")
    _write_sheet_from_rows(ws_bib, bib_rows, "Referencias detectadas")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
